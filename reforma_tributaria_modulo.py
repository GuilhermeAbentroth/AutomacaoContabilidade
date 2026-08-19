import customtkinter as ctk
import tkinter as tk
from tkinter import scrolledtext, filedialog, messagebox
import threading
import time
import re
import os
import unicodedata
from datetime import datetime

import requests
import openpyxl
from utils import isolar_scroll_mouse


class ReformaTributariaModulo:
    """
    Reforma Tributária — versão automatizada do módulo Societário para lotes
    de empresas: recebe uma planilha Excel (CNPJ + Razão Social) e consulta,
    empresa por empresa, apenas se ela é ou não optante do Simples Nacional
    (publica.cnpj.ws como fonte primária, com fallback em ReceitaWS e
    BrasilAPI quando a primeira falha ou atinge o limite de requisições),
    devolvendo a própria planilha preenchida na coluna C.
    """

    API_URL = "https://publica.cnpj.ws/cnpj/{cnpj}"
    # publica.cnpj.ws libera só ~3 requisições/minuto no plano público — abaixo
    # disso o lote inteiro cai em 429 a partir da 4ª empresa. 20s mantém a
    # margem de segurança (3/min) mesmo com o overhead de rede.
    DELAY_ENTRE_CONSULTAS = 20.0

    COL_CNPJ_NOMES = {"cnpj"}
    COL_RAZAO_NOMES = {"razao social", "razão social", "nome", "empresa"}

    def __init__(self, parent):
        self.parent = parent
        self._parar = False
        self._processando = False
        self._caminho_resultado = None

    # ------------------------------------------------------------------
    # UI
    # ------------------------------------------------------------------
    def tela_reforma_tributaria(self):
        self.parent.limpar_tela()

        f_head = ctk.CTkFrame(self.parent.container, fg_color="transparent")
        f_head.pack(fill="x", pady=(8, 4))
        ctk.CTkLabel(f_head, text="⚖️ Reforma Tributária — Consulta em Lote (Simples Nacional)",
                     font=("Arial", 20, "bold"), text_color="#1f538d").pack()
        ctk.CTkLabel(f_head,
                     text="Lê uma planilha com CNPJ + Razão Social e devolve o mesmo arquivo "
                          "preenchido na coluna C com a situação de cada empresa no Simples Nacional.",
                     font=("Arial", 11), text_color="gray").pack()
        ctk.CTkLabel(f_head,
                     text="⏳ ~20s por empresa (limite de requisições das APIs públicas de CNPJ) — "
                          "para 20 empresas, conte uns 6-7 minutos.",
                     font=("Arial", 10), text_color="#e67e22").pack()

        f_card = ctk.CTkFrame(self.parent.container)
        f_card.pack(fill="x", padx=40, pady=10)

        f_excel = ctk.CTkFrame(f_card, fg_color="transparent")
        f_excel.pack(fill="x", padx=20, pady=(20, 10))

        ctk.CTkLabel(f_excel, text="Planilha (CNPJ / Razão Social):",
                     font=("Arial", 12, "bold"), anchor="w").pack(side="left")

        self.ent_excel = ctk.CTkEntry(f_excel, font=("Arial", 12), width=420)
        self.ent_excel.pack(side="left", padx=10)
        ultimo = self.parent.configuracoes.get("ultimo_excel_reforma", "")
        if ultimo:
            self.ent_excel.insert(0, ultimo)

        ctk.CTkButton(f_excel, text="Procurar", width=100,
                      command=self._selecionar_excel).pack(side="left", padx=5)

        f_status = ctk.CTkFrame(f_card, fg_color="transparent")
        f_status.pack(fill="x", padx=20, pady=(0, 16))
        self.lbl_status = ctk.CTkLabel(f_status, text="", font=("Arial", 11), text_color="gray")
        self.lbl_status.pack(side="left")
        self.lbl_contador = ctk.CTkLabel(f_status, text="", font=("Arial", 11, "bold"),
                                          text_color="#1f538d")
        self.lbl_contador.pack(side="right")

        # Log — mesmo padrão (scrolledtext + parent.log_msg) usado no resto do sistema
        f_log_area = ctk.CTkFrame(self.parent.container, fg_color="transparent")
        f_log_area.pack(fill="both", expand=True, padx=40, pady=5)
        self.parent.txt_log = scrolledtext.ScrolledText(
            f_log_area, width=125, height=16, state='disabled',
            bg="#1e1e1e", fg="white", font=("Consolas", 10))
        self.parent.txt_log.pack(fill="both", expand=True)
        isolar_scroll_mouse(self.parent.txt_log)

        f_btns = ctk.CTkFrame(self.parent.container, fg_color="transparent")
        f_btns.pack(pady=16)

        self.btn_iniciar = ctk.CTkButton(
            f_btns, text="▶ Iniciar Consulta em Lote", width=240, height=44,
            font=("Arial", 13, "bold"), fg_color="#27ae60", hover_color="#1e8449",
            command=self._iniciar_lote)
        self.btn_iniciar.pack(side="left", padx=6)

        self.btn_parar = ctk.CTkButton(
            f_btns, text="🛑 Parar", width=120, height=44,
            fg_color="#c0392b", hover_color="#962d22",
            command=self._parar_lote, state="disabled")
        self.btn_parar.pack(side="left", padx=6)

        self.btn_abrir = ctk.CTkButton(
            f_btns, text="📂 Abrir Arquivo Gerado", width=200, height=44,
            fg_color="#2980b9", hover_color="#1f618d",
            command=self._abrir_resultado, state="disabled")
        self.btn_abrir.pack(side="left", padx=6)

        ctk.CTkButton(f_btns, text="Voltar", width=110, height=44,
                      fg_color="gray", hover_color="darkgray",
                      command=self.parent.mostrar_menu_inicial).pack(side="left", padx=6)

    # ------------------------------------------------------------------
    # Seleção de arquivo
    # ------------------------------------------------------------------
    def _selecionar_excel(self):
        caminho = filedialog.askopenfilename(
            title="Selecione a planilha de empresas",
            filetypes=[("Excel", "*.xlsx *.xls"), ("Todos os arquivos", "*.*")])
        if caminho:
            self.ent_excel.delete(0, tk.END)
            self.ent_excel.insert(0, caminho)

    # ------------------------------------------------------------------
    # Controle do lote
    # ------------------------------------------------------------------
    def _iniciar_lote(self):
        if self._processando:
            return
        caminho = self.ent_excel.get().strip()
        if not caminho or not os.path.exists(caminho):
            messagebox.showwarning("Planilha inválida", "Selecione uma planilha Excel válida.")
            return

        self.parent.salvar_config("ultimo_excel_reforma", caminho)
        self._parar = False
        self._processando = True
        self._caminho_resultado = None
        self.btn_iniciar.configure(state="disabled")
        self.btn_parar.configure(state="normal")
        self.btn_abrir.configure(state="disabled")
        self.lbl_status.configure(text="Processando...", text_color="#e67e22")

        threading.Thread(target=self._processar_lote, args=(caminho,), daemon=True).start()

    def _parar_lote(self):
        self._parar = True
        self.parent.log_msg("Solicitação de parada recebida — encerra após a empresa atual.", "erro")

    def _abrir_resultado(self):
        if self._caminho_resultado and os.path.exists(self._caminho_resultado):
            os.startfile(self._caminho_resultado)

    # ------------------------------------------------------------------
    # Localização das colunas de entrada
    # ------------------------------------------------------------------
    def _normalizar(self, texto: str) -> str:
        texto = unicodedata.normalize('NFD', texto)
        texto = ''.join(c for c in texto if unicodedata.category(c) != 'Mn')
        return texto.strip().lower()

    def _localizar_colunas(self, planilha):
        """Procura as colunas de CNPJ e Razão Social pelo cabeçalho (linha 1);
        se não encontrar, assume A=CNPJ e B=Razão Social (layout padrão)."""
        col_cnpj, col_razao = 1, 2
        achou_cnpj = achou_razao = False
        for cel in planilha[1]:
            if cel.value is None:
                continue
            titulo = self._normalizar(str(cel.value))
            if not achou_cnpj and titulo in self.COL_CNPJ_NOMES:
                col_cnpj = cel.column
                achou_cnpj = True
            elif not achou_razao and titulo in self.COL_RAZAO_NOMES:
                col_razao = cel.column
                achou_razao = True
        return col_cnpj, col_razao

    # ------------------------------------------------------------------
    # Processamento em lote
    # ------------------------------------------------------------------
    def _processar_lote(self, caminho_excel):
        try:
            wb = openpyxl.load_workbook(caminho_excel)
            planilha = wb.active
        except Exception as e:
            self.parent.log_msg(f"Erro ao abrir a planilha: {e}", "erro")
            self._finalizar_lote()
            return

        col_cnpj, col_razao = self._localizar_colunas(planilha)
        col_resultado = 3  # Coluna C, conforme solicitado

        # Cabeçalho da coluna C (só grava se estiver vazia, pra não sobrescrever algo já existente)
        cel_header = planilha.cell(row=1, column=col_resultado)
        if not cel_header.value:
            cel_header.value = "SITUAÇÃO SIMPLES NACIONAL"

        linhas = []
        for r in range(2, planilha.max_row + 1):
            valor_cnpj = planilha.cell(row=r, column=col_cnpj).value
            if valor_cnpj is not None and str(valor_cnpj).strip() != "":
                linhas.append(r)

        total = len(linhas)
        if total == 0:
            self.parent.log_msg("Nenhuma empresa encontrada na planilha (coluna CNPJ vazia).", "erro")
            self._finalizar_lote()
            return

        self.parent.log_msg(
            f"Iniciando consulta em lote — {total} empresa(s) — Simples Nacional...",
            "info", divisor=True)

        qtd_optante = qtd_nao_optante = qtd_erro = 0

        for idx, num_linha in enumerate(linhas, start=1):
            if self._parar:
                self.parent.log_msg("Lote interrompido pelo usuário.", "erro")
                break

            self.parent.root.after(0, lambda i=idx, t=total: self.lbl_contador.configure(
                text=f"{i}/{t} empresas"))

            cnpj_bruto = str(planilha.cell(row=num_linha, column=col_cnpj).value or "")
            razao_bruta = str(planilha.cell(row=num_linha, column=col_razao).value or "").strip()
            cnpj = self._limpar_cnpj(cnpj_bruto)
            cnpj_fmt = self._fmt_cnpj(cnpj) if len(cnpj) == 14 else cnpj_bruto

            self.parent.log_msg(f"Empresa {idx}/{total}: {cnpj_fmt} {razao_bruta}", "info")

            resultado = self._consultar_situacao_sn(cnpj)

            if resultado["optante"] is True:
                texto_resultado = "Optante Simples Nacional"
                qtd_optante += 1
                self.parent.log_msg(
                    f"    ✅ Optante Simples Nacional  (fonte: {resultado['fonte']})", "sucesso")
            elif resultado["optante"] is False:
                texto_resultado = "Não Optante Simples Nacional"
                qtd_nao_optante += 1
                self.parent.log_msg(
                    f"    ⬜ Não Optante Simples Nacional  (fonte: {resultado['fonte']})", "info")
            else:
                texto_resultado = f"Erro na consulta: {resultado.get('erro', 'indefinido')}"
                qtd_erro += 1
                self.parent.log_msg(
                    f"    ⚠ Não foi possível determinar — {resultado.get('erro', '')}", "erro")

            planilha.cell(row=num_linha, column=col_resultado).value = texto_resultado

            if idx < total and not self._parar:
                time.sleep(self.DELAY_ENTRE_CONSULTAS)

        pasta_saida = os.path.join(self.parent.pasta_exe, "saida de arquivos")
        os.makedirs(pasta_saida, exist_ok=True)
        nome_base = os.path.splitext(os.path.basename(caminho_excel))[0]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        caminho_saida = os.path.join(pasta_saida, f"{nome_base}_REFORMA_TRIBUTARIA_{timestamp}.xlsx")

        try:
            wb.save(caminho_saida)
            self._caminho_resultado = caminho_saida
            self.parent.log_msg(
                f"Concluído — {qtd_optante} optante(s), {qtd_nao_optante} não optante(s), "
                f"{qtd_erro} com erro. Arquivo salvo em: {caminho_saida}", "sucesso", divisor=True)
            self.parent.root.after(0, lambda: self.btn_abrir.configure(state="normal"))
        except Exception as e:
            self.parent.log_msg(f"Erro ao salvar planilha de resultado: {e}", "erro")

        self._finalizar_lote()

    def _finalizar_lote(self):
        self._processando = False
        interrompido = self._parar
        self.parent.root.after(0, lambda: self.btn_iniciar.configure(state="normal"))
        self.parent.root.after(0, lambda: self.btn_parar.configure(state="disabled"))
        self.parent.root.after(0, lambda: self.lbl_status.configure(
            text="Interrompido." if interrompido else "Concluído.",
            text_color="#c0392b" if interrompido else "#27ae60"))

    # ------------------------------------------------------------------
    # Consulta — apenas "é ou não optante do Simples Nacional"
    # ------------------------------------------------------------------
    def _limpar_cnpj(self, texto: str) -> str:
        return re.sub(r'[^A-Za-z0-9]', '', str(texto)).upper()[:14]

    def _fmt_cnpj(self, d: str) -> str:
        return f"{d[:2]}.{d[2:5]}.{d[5:8]}/{d[8:12]}-{d[12:]}"

    def _tempo_espera_429(self, resp) -> int:
        """Respeita o header Retry-After quando a API informa; senão espera
        passar de uma janela cheia de 1 minuto (limite costuma ser por minuto)."""
        retry_after = resp.headers.get("Retry-After")
        if retry_after:
            try:
                return max(1, min(int(float(retry_after)), 90))
            except ValueError:
                pass
        return 65

    def _consultar_situacao_sn(self, cnpj: str) -> dict:
        """Tenta, em ordem, publica.cnpj.ws → ReceitaWS → BrasilAPI — todas
        fontes públicas e gratuitas que expõem só "optante Simples: Sim/Não".
        Cada uma tenta 1x, e em caso de 429 espera e tenta mais 1x antes de
        passar pra próxima fonte, pra não derrubar o lote inteiro."""
        if len(cnpj) != 14:
            return {"optante": None, "fonte": None, "erro": "CNPJ inválido"}

        for consultar in (self._consultar_publica_cnpj_ws,
                           self._consultar_receitaws,
                           self._consultar_brasilapi):
            resultado = consultar(cnpj)
            if resultado is not None:
                return resultado

        return {"optante": None, "fonte": None, "erro": "Todas as fontes falharam"}

    def _consultar_publica_cnpj_ws(self, cnpj: str):
        """Fonte primária — dados oficiais da Receita Federal via publica.cnpj.ws."""
        for tentativa in range(2):
            try:
                resp = requests.get(self.API_URL.format(cnpj=cnpj), timeout=15,
                                     headers={"Accept": "application/json"})
            except requests.exceptions.RequestException:
                return None

            if resp.status_code == 200:
                simples = resp.json().get("simples")
                if not simples:
                    # Sem registro de Simples nessa fonte (comum em entidades sem
                    # fins lucrativos, associações etc., que nunca podem optar) —
                    # a própria Receita Federal não retorna esse bloco. É uma
                    # resposta válida e definitiva: não é optante.
                    return {"optante": False, "fonte": "Receita Federal (publica.cnpj.ws)", "erro": None}
                valor = simples.get("simples")
                if valor is None:
                    return None  # bloco existe mas sem o campo esperado — tenta a próxima fonte
                optante = str(valor).strip().lower() == "sim"
                return {"optante": optante, "fonte": "Receita Federal (publica.cnpj.ws)", "erro": None}

            if resp.status_code == 404:
                return {"optante": None, "fonte": None, "erro": "CNPJ não encontrado"}

            if resp.status_code == 429 and tentativa == 0:
                espera = self._tempo_espera_429(resp)
                self.parent.log_msg(
                    f"    ⏳ Limite de requisições (publica.cnpj.ws) — aguardando {espera}s...", "info")
                time.sleep(espera)
                continue

            return None  # outro erro — tenta a próxima fonte

        return None

    def _consultar_receitaws(self, cnpj: str):
        for tentativa in range(2):
            try:
                resp = requests.get(f"https://receitaws.com.br/v1/cnpj/{cnpj}", timeout=10,
                                     headers={"Accept": "application/json"})
            except requests.exceptions.RequestException:
                return None

            if resp.status_code == 200:
                sn_str = str(resp.json().get("simples") or "").strip().upper().rstrip(".")
                if sn_str in ("SIM", "NÃO", "NAO"):
                    return {"optante": sn_str == "SIM", "fonte": "ReceitaWS", "erro": None}
                return None

            if resp.status_code == 429 and tentativa == 0:
                espera = self._tempo_espera_429(resp)
                self.parent.log_msg(
                    f"    ⏳ Limite de requisições (ReceitaWS) — aguardando {espera}s...", "info")
                time.sleep(espera)
                continue

            return None

        return None

    def _consultar_brasilapi(self, cnpj: str):
        for tentativa in range(2):
            try:
                resp = requests.get(f"https://brasilapi.com.br/api/cnpj/v1/{cnpj}", timeout=10,
                                     headers={"Accept": "application/json"})
            except requests.exceptions.RequestException:
                return None

            if resp.status_code == 200:
                sn = resp.json().get("opcao_pelo_simples")
                if sn is None:
                    return None
                return {"optante": bool(sn), "fonte": "BrasilAPI", "erro": None}

            if resp.status_code == 429 and tentativa == 0:
                espera = self._tempo_espera_429(resp)
                self.parent.log_msg(
                    f"    ⏳ Limite de requisições (BrasilAPI) — aguardando {espera}s...", "info")
                time.sleep(espera)
                continue

            return None

        return None
